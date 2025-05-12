import os
import pandas as pd
from typing import Dict, List, Any, Optional
from fastapi import UploadFile, HTTPException, status
from openpyxl import load_workbook
from pathlib import Path

from app.core.config import settings
from app.utils.file_utils import save_upload_file


class SpreadsheetService:
    """Service for handling Excel spreadsheet operations"""
    
    @staticmethod
    async def upload_spreadsheet(file: UploadFile, user_id: str) -> Dict[str, Any]:
        """
        Upload and save a spreadsheet file
        
        Args:
            file: The uploaded file
            user_id: The ID of the user uploading the file
            
        Returns:
            Dict with file metadata
        """
        # Check file extension
        file_ext = file.filename.split('.')[-1].lower()
        if file_ext not in settings.ALLOWED_EXTENSIONS:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"File type not supported. Allowed types: {', '.join(settings.ALLOWED_EXTENSIONS)}"
            )
        
        # Create user upload directory if it doesn't exist
        user_upload_dir = Path(settings.UPLOAD_DIR) / user_id
        os.makedirs(user_upload_dir, exist_ok=True)
        
        # Save the uploaded file
        file_path = await save_upload_file(file, user_upload_dir)
        
        # Get basic metadata
        metadata = {
            "filename": file.filename,
            "content_type": file.content_type,
            "size": os.path.getsize(file_path),
            "path": str(file_path),
            "user_id": user_id,
        }
        
        # Add sheet names if Excel file
        if file_ext in ["xlsx", "xls"]:
            try:
                workbook = load_workbook(file_path, read_only=True)
                metadata["sheets"] = workbook.sheetnames
            except Exception as e:
                os.remove(file_path)
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Invalid Excel file: {str(e)}"
                )
        
        return metadata
    
    @staticmethod
    def read_spreadsheet(file_path: str, sheet_name: Optional[str] = None) -> pd.DataFrame:
        """
        Read a spreadsheet into a pandas DataFrame
        
        Args:
            file_path: Path to the spreadsheet file
            sheet_name: Name of the sheet to read (only for Excel files)
            
        Returns:
            pandas DataFrame with the spreadsheet data
        """
        if not os.path.exists(file_path):
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="File not found"
            )
        
        file_ext = file_path.split('.')[-1].lower()
        
        try:
            if file_ext == "csv":
                return pd.read_csv(file_path)
            elif file_ext in ["xlsx", "xls"]:
                return pd.read_excel(file_path, sheet_name=sheet_name)
            else:
                raise HTTPException(
                    status_code=status.HTTP_400_BAD_REQUEST,
                    detail=f"Unsupported file type: {file_ext}"
                )
        except Exception as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Error reading file: {str(e)}"
            )
    
    @staticmethod
    def generate_financial_scenario(
        data: pd.DataFrame, 
        parameters: Dict[str, Any]
    ) -> Dict[str, Any]:
        """
        Generate a financial scenario based on input data and parameters
        
        Args:
            data: DataFrame with financial data
            parameters: Dictionary with scenario parameters
            
        Returns:
            Dictionary with generated scenario data
        """
        # This would contain the actual financial modeling logic
        # For now, we'll just return a placeholder
        
        return {
            "scenario_name": parameters.get("name", "New Scenario"),
            "description": parameters.get("description", ""),
            "data_shape": data.shape,
            "columns": data.columns.tolist(),
            "sample_data": data.head(5).to_dict(orient="records"),
            "parameters": parameters,
        } 