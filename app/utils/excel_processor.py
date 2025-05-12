"""
Excel Processor para o Habitus Forecast.

Este módulo contém funcionalidades para processamento, validação e preparação
de dados financeiros a partir de planilhas Excel.
"""

import os
import pandas as pd
import numpy as np
import io
from typing import Dict, List, Any, Optional, Union, Tuple, Set, Callable
from pathlib import Path
import logging
from enum import Enum
from datetime import datetime
from fastapi import HTTPException, status
from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException

# Configurar logger
logger = logging.getLogger(__name__)


class ExcelValidationError(Exception):
    """Exceção para erro de validação em arquivos Excel."""
    pass


class FinancialCategory(Enum):
    """Categorias financeiras para classificação dos dados."""
    REVENUE = "receitas"
    VARIABLE_COSTS = "custos_variaveis"
    PERSONNEL_EXPENSES = "despesas_pessoal"
    COMMERCIAL_EXPENSES = "despesas_comerciais"
    ADMIN_EXPENSES = "despesas_administrativas"
    INVESTMENTS = "investimentos"
    CONTRIBUTION_MARGIN = "margem_contribuicao"
    CASH_FLOW = "fluxo_de_caixa"
    FINAL_BALANCE = "saldo_final"


class ExcelProcessor:
    """
    Processador de arquivos Excel para dados financeiros.
    
    Esta classe é responsável por ler, validar e processar dados financeiros
    de arquivos Excel, extraindo categorias específicas como receitas, custos
    e despesas.
    
    Attributes:
        financial_data (Dict[str, pd.DataFrame]): Dados financeiros extraídos.
        metadata (Dict[str, Any]): Metadados do arquivo processado.
    """
    
    # Mapeamento de palavras-chave para identificar categorias
    CATEGORY_KEYWORDS = {
        FinancialCategory.REVENUE: ["receita", "faturamento", "vendas", "entrada"],
        FinancialCategory.VARIABLE_COSTS: ["custo variável", "custo operacional", "matéria prima"],
        FinancialCategory.PERSONNEL_EXPENSES: ["despesa pessoal", "salário", "remuneração", "benefício"],
        FinancialCategory.COMMERCIAL_EXPENSES: ["despesa comercial", "marketing", "vendas", "comissão"],
        FinancialCategory.ADMIN_EXPENSES: ["despesa administrativa", "escritório", "aluguel", "utilidade"],
        FinancialCategory.INVESTMENTS: ["investimento", "aquisição", "ativo", "imobilizado"]
    }
    
    def __init__(self):
        """Inicializa o processador de Excel."""
        self.financial_data: Dict[str, pd.DataFrame] = {}
        self.metadata: Dict[str, Any] = {
            "processing_date": datetime.now(),
            "sheet_names": [],
            "categories_found": []
        }
    
    def process_excel_data(self, content: bytes) -> Dict[str, Any]:
        """
        Processa um arquivo Excel contendo dados financeiros.
        
        Args:
            content: Conteúdo do arquivo Excel em bytes.
            
        Returns:
            Dicionário com dados financeiros processados e metadados.
            
        Raises:
            ExcelValidationError: Se o arquivo não atender os requisitos.
        """
        # Ler o arquivo Excel em memória
        try:
            excel_file = io.BytesIO(content)
            # Ler todas as planilhas
            excel_data = pd.read_excel(excel_file, sheet_name=None)
        except Exception as e:
            logger.error(f"Erro ao ler arquivo Excel: {str(e)}")
            raise ExcelValidationError(f"Erro ao ler arquivo Excel: {str(e)}")
        
        # Atualizar metadados
        self.metadata["sheet_names"] = list(excel_data.keys())
        self.metadata["total_sheets"] = len(excel_data)
        
        # Validar a estrutura do arquivo
        self._validate_excel_structure(excel_data)
        
        # Extrair dados financeiros
        self._extract_financial_data(excel_data)
        
        # Validar os dados extraídos
        self._validate_financial_data()
        
        # Preparar a resposta
        response = {
            "status": "success",
            "message": "Dados financeiros processados com sucesso",
            "categories": [
                {"id": cat.value, "name": self._format_category_name(cat.value), 
                 "description": self._get_category_description(cat)}
                for cat in self.metadata.get("categories_found", [])
            ],
            "data": self.financial_data,
            "metadata": self.metadata
        }
        
        return response
    
    def _validate_excel_structure(self, excel_data: Dict[str, pd.DataFrame]) -> None:
        """
        Valida a estrutura do arquivo Excel.
        
        Args:
            excel_data: Dicionário com os dados das planilhas.
            
        Raises:
            ExcelValidationError: Se a estrutura não for válida.
        """
        if not excel_data:
            raise ExcelValidationError("O arquivo Excel não contém planilhas.")
        
        # Verifica se pelo menos uma planilha tem dados suficientes
        has_valid_data = False
        for sheet_name, df in excel_data.items():
            if not df.empty and df.shape[0] > 0 and df.shape[1] > 0:
                has_valid_data = True
                break
        
        if not has_valid_data:
            raise ExcelValidationError("O arquivo Excel não contém dados válidos em nenhuma planilha.")
    
    def _extract_financial_data(self, excel_data: Dict[str, pd.DataFrame]) -> None:
        """
        Extrai dados financeiros do arquivo Excel.
        
        Args:
            excel_data: Dicionário com os dados das planilhas.
        """
        # Inicializar contador de categorias encontradas
        categories_found = set()
        
        # Processar cada planilha
        for sheet_name, df in excel_data.items():
            if df.empty:
                continue
            
            # Tenta identificar a categoria da planilha
            category = self._identify_sheet_category(sheet_name, df)
            
            if category:
                # Processar dados conforme a categoria
                processed_df = self._process_dataframe_by_category(df, category)
                
                # Armazenar dados processados
                self.financial_data[category.value] = processed_df
                categories_found.add(category)
                
                # Registrar no log
                logger.info(f"Planilha '{sheet_name}' processada como '{category.value}'")
        
        # Atualizar metadados
        self.metadata["categories_found"] = list(categories_found)
        self.metadata["total_categories"] = len(categories_found)
    
    def _identify_sheet_category(
        self, 
        sheet_name: str, 
        df: pd.DataFrame
    ) -> Optional[FinancialCategory]:
        """
        Identifica a categoria financeira de uma planilha.
        
        Args:
            sheet_name: Nome da planilha.
            df: DataFrame da planilha.
            
        Returns:
            Categoria financeira identificada ou None.
        """
        # Converter para minúsculas para facilitar as comparações
        sheet_name_lower = sheet_name.lower()
        
        # Verificar se o nome da planilha contém uma palavra-chave de categoria
        for category, keywords in self.CATEGORY_KEYWORDS.items():
            if any(keyword in sheet_name_lower for keyword in keywords):
                return category
        
        # Se não encontrou pelo nome da planilha, tenta pelo conteúdo da planilha
        # Converte todas as colunas de texto para string e concatena
        text_content = ""
        for col in df.columns:
            if df[col].dtype == 'object':
                text_content += " ".join(df[col].astype(str).tolist())
                text_content = text_content.lower()
        
        # Verifica se o conteúdo contém palavras-chave
        for category, keywords in self.CATEGORY_KEYWORDS.items():
            if any(keyword in text_content for keyword in keywords):
                return category
        
        # Tenta inferir pelo nome das colunas
        columns_text = " ".join(df.columns.astype(str)).lower()
        for category, keywords in self.CATEGORY_KEYWORDS.items():
            if any(keyword in columns_text for keyword in keywords):
                return category
        
        return None
    
    def _process_dataframe_by_category(
        self, 
        df: pd.DataFrame, 
        category: FinancialCategory
    ) -> pd.DataFrame:
        """
        Processa um DataFrame conforme sua categoria.
        
        Args:
            df: DataFrame a ser processado.
            category: Categoria financeira.
            
        Returns:
            DataFrame processado.
        """
        # Cópia para evitar modificações no original
        processed_df = df.copy()
        
        # Remove linhas totalmente vazias
        processed_df = processed_df.dropna(how='all')
        
        # Tenta identificar e padronizar colunas de data/período
        date_columns = [col for col in processed_df.columns if 
                         'data' in str(col).lower() or 
                         'período' in str(col).lower() or 
                         'mes' in str(col).lower() or
                         'mês' in str(col).lower()]
        
        if date_columns:
            # Padronizar como datetime se possível
            for col in date_columns:
                try:
                    processed_df[col] = pd.to_datetime(processed_df[col], errors='coerce')
                except:
                    pass  # Mantém como está se não for possível converter
        
        # Padronizar colunas numéricas (converte para float)
        numeric_columns = processed_df.select_dtypes(include=['number']).columns
        for col in numeric_columns:
            processed_df[col] = processed_df[col].astype(float)
        
        return processed_df
    
    def _validate_financial_data(self) -> None:
        """
        Valida os dados financeiros extraídos.
        
        Raises:
            ExcelValidationError: Se os dados não forem válidos.
        """
        if not self.financial_data:
            raise ExcelValidationError(
                "Não foi possível extrair dados financeiros. "
                "Verifique se o arquivo contém categorias reconhecíveis."
            )
        
        # Verificações adicionais para garantir dados coerentes
        for category, df in self.financial_data.items():
            # Verificar se há colunas numéricas
            numeric_columns = df.select_dtypes(include=['number']).columns
            if len(numeric_columns) == 0:
                raise ExcelValidationError(
                    f"A categoria '{category}' não contém colunas numéricas para análise."
                )
    
    def _format_category_name(self, category_name: str) -> str:
        """
        Formata o nome da categoria para apresentação.
        
        Args:
            category_name: Nome da categoria.
            
        Returns:
            Nome formatado.
        """
        return " ".join(word.capitalize() for word in category_name.replace("_", " ").split())
    
    def _get_category_description(self, category: FinancialCategory) -> str:
        """
        Retorna uma descrição para a categoria.
        
        Args:
            category: Categoria financeira.
            
        Returns:
            Descrição da categoria.
        """
        descriptions = {
            FinancialCategory.REVENUE: "Entradas de recursos financeiros",
            FinancialCategory.VARIABLE_COSTS: "Custos diretos relacionados à operação",
            FinancialCategory.PERSONNEL_EXPENSES: "Despesas com pessoal e folha de pagamento",
            FinancialCategory.COMMERCIAL_EXPENSES: "Despesas comerciais e de marketing",
            FinancialCategory.ADMIN_EXPENSES: "Despesas administrativas e gerais",
            FinancialCategory.INVESTMENTS: "Aplicações de capital",
            FinancialCategory.CONTRIBUTION_MARGIN: "Margem de contribuição financeira",
            FinancialCategory.CASH_FLOW: "Fluxo de caixa operacional",
            FinancialCategory.FINAL_BALANCE: "Saldo final acumulado"
        }
        
        return descriptions.get(category, "Categoria financeira")
    
    def analyze_trends(self) -> Dict[str, Any]:
        """
        Analisa tendências nos dados financeiros.
        
        Returns:
            Dicionário com análises de tendências.
        """
        if not self.financial_data:
            return {"status": "error", "message": "Sem dados para análise"}
        
        trends = {}
        
        # Para cada categoria, calcula tendências básicas
        for category, df in self.financial_data.items():
            numeric_cols = df.select_dtypes(include=['number']).columns
            
            if len(numeric_cols) < 2:
                continue
            
            # Calcular estatísticas básicas
            category_trends = {
                "media": df[numeric_cols].mean().to_dict(),
                "mediana": df[numeric_cols].median().to_dict(),
                "tendencia": {},
                "crescimento": {}
            }
            
            # Calcular tendência (crescente, decrescente ou estável)
            for col in numeric_cols:
                values = df[col].dropna().values
                if len(values) < 2:
                    continue
                
                # Calcular taxa de crescimento
                growth_rate = (values[-1] / values[0]) - 1 if values[0] != 0 else 0
                category_trends["crescimento"][col] = growth_rate
                
                # Determinar tendência
                if growth_rate > 0.05:  # Crescimento de mais de 5%
                    trend = "crescente"
                elif growth_rate < -0.05:  # Queda de mais de 5%
                    trend = "decrescente"
                else:
                    trend = "estável"
                
                category_trends["tendencia"][col] = trend
            
            trends[category] = category_trends
        
        return {
            "status": "success",
            "message": "Análise de tendências concluída",
            "trends": trends
        }

    def _clean_dataframe(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Realiza limpeza básica em um DataFrame.
        
        Args:
            df: DataFrame a ser limpo.
            
        Returns:
            DataFrame limpo.
        """
        # Substitui strings vazias por NaN
        df = df.replace(r'^\s*$', np.nan, regex=True)
        
        # Remove colunas completamente vazias
        df = df.dropna(axis=1, how='all')
        
        # Converte colunas numéricas para o tipo correto
        for col in df.columns:
            if pd.api.types.is_numeric_dtype(df[col]) or self._is_numeric_column(df[col]):
                try:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
                except Exception:
                    pass
        
        return df
    
    def _is_numeric_column(self, column: pd.Series) -> bool:
        """
        Verifica se uma coluna contém valores numéricos.
        
        Args:
            column: Série pandas a ser verificada.
            
        Returns:
            True se a coluna for numérica, False caso contrário.
        """
        # Remove valores nulos para teste
        non_null = column.dropna()
        if non_null.empty:
            return False
        
        # Se já for numérico, retorna True
        if pd.api.types.is_numeric_dtype(non_null):
            return True
        
        # Tenta converter para numérico
        try:
            pd.to_numeric(non_null)
            return True
        except (ValueError, TypeError):
            return False
    
    def validate_structure(self, required_categories: Optional[List[FinancialCategory]] = None) -> bool:
        """
        Valida a estrutura da planilha para confirmar se contém as categorias financeiras necessárias.
        
        Args:
            required_categories: Lista de categorias que devem estar presentes.
                Se None, verifica todas as categorias em FinancialCategory.
                
        Returns:
            True se a estrutura for válida, False caso contrário.
            
        Raises:
            ValueError: Se o Excel não foi lido antes da validação.
        """
        if not self.financial_data:
            raise ValueError("Excel não foi lido. Execute process_excel_data() primeiro.")
        
        if required_categories is None:
            required_categories = list(FinancialCategory)
        
        found_categories = set()
        missing_categories = []
        
        # Procura as categorias em todas as abas
        for category in required_categories:
            if category.value in self.financial_data:
                found_categories.add(category)
        
        # Identifica categorias ausentes
        for category in required_categories:
            if category not in found_categories:
                missing_categories.append(category.value)
                self.metadata["validation_errors"].append({
                    "type": "missing_category",
                    "category": category.value
                })
        
        is_valid = len(missing_categories) == 0
        self.metadata["is_valid_structure"] = is_valid
        self.metadata["missing_categories"] = missing_categories
        
        return is_valid
    
    def extract_financial_data(self) -> Dict[str, pd.DataFrame]:
        """
        Extrai os dados financeiros das categorias encontradas nas planilhas.
        
        Returns:
            Dicionário com DataFrames para cada categoria financeira encontrada.
            
        Raises:
            ValueError: Se o Excel não foi lido antes da extração.
        """
        if not self.financial_data:
            raise ValueError("Excel não foi lido. Execute process_excel_data() primeiro.")
        
        return self.financial_data
    
    def _generate_data_summary(self) -> None:
        """
        Gera um resumo dos dados financeiros extraídos.
        
        Esta função calcula estatísticas básicas para cada categoria financeira
        e armazena os resultados no atributo metadata.
        """
        for category, data in self.financial_data.items():
            numeric_columns = data.select_dtypes(include=['number']).columns.tolist()
            
            summary = {
                "row_count": len(data),
                "column_count": len(data.columns),
                "numeric_columns": numeric_columns,
                "missing_data_percentage": data.isna().mean().mean() * 100
            }
            
            # Adiciona totais para colunas numéricas
            if numeric_columns:
                column_sums = {}
                for col in numeric_columns:
                    try:
                        column_sums[col] = data[col].sum()
                    except:
                        pass
                summary["column_sums"] = column_sums
            
            self.metadata["data_summary"][category] = summary
    
    def export_processed_data(self, output_path: Union[str, Path]) -> Dict[str, str]:
        """
        Exporta os dados processados para um arquivo Excel.
        
        Args:
            output_path: Caminho onde o arquivo de saída será salvo.
            
        Returns:
            Dicionário com informações sobre a exportação.
            
        Raises:
            ValueError: Se os dados financeiros não foram extraídos.
        """
        if not self.financial_data:
            raise ValueError("Dados financeiros não extraídos. Execute process_excel_data() primeiro.")
        
        output_path = Path(output_path) if isinstance(output_path, str) else output_path
        
        # Cria o diretório se não existir
        output_path.parent.mkdir(parents=True, exist_ok=True)
        
        try:
            with pd.ExcelWriter(output_path) as writer:
                # Exporta cada categoria em uma aba separada
                for category, df in self.financial_data.items():
                    sheet_name = category[:31]  # Limite de 31 caracteres para nomes de abas no Excel
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                
                # Adiciona uma aba com metadados
                metadata_df = pd.DataFrame([
                    ["Arquivo original", self.file_path.name],
                    ["Data de processamento", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
                    ["Tamanho do arquivo", f"{self.metadata['file_size'] / 1024:.2f} KB"],
                    ["Categorias encontradas", ", ".join(self.financial_data.keys())]
                ], columns=["Propriedade", "Valor"])
                
                metadata_df.to_excel(writer, sheet_name="Metadados", index=False)
            
            return {
                "success": True,
                "output_path": str(output_path),
                "file_size": output_path.stat().st_size,
                "categories_exported": list(self.financial_data.keys())
            }
            
        except Exception as e:
            error_msg = f"Erro ao exportar dados: {str(e)}"
            logger.error(error_msg)
            raise RuntimeError(error_msg)
    
    @staticmethod
    def check_file_integrity(file_path: Union[str, Path]) -> Dict[str, Any]:
        """
        Verifica a integridade de um arquivo Excel sem carregá-lo completamente.
        
        Args:
            file_path: Caminho para o arquivo Excel a ser verificado.
            
        Returns:
            Dicionário com informações sobre a integridade do arquivo.
        """
        file_path = Path(file_path) if isinstance(file_path, str) else file_path
        
        integrity_check = {
            "file_exists": False,
            "is_excel_file": False,
            "file_size": 0,
            "is_corrupt": False,
            "can_open": False,
            "sheet_count": 0,
            "error_message": None
        }
        
        if not file_path.exists():
            integrity_check["error_message"] = "Arquivo não encontrado"
            return integrity_check
        
        integrity_check["file_exists"] = True
        integrity_check["file_size"] = file_path.stat().st_size
        
        if file_path.suffix.lower() not in ['.xlsx', '.xls']:
            integrity_check["error_message"] = f"Formato de arquivo inválido: {file_path.suffix}"
            return integrity_check
        
        integrity_check["is_excel_file"] = True
        
        try:
            # Tenta abrir o arquivo com openpyxl para verificar a integridade
            workbook = load_workbook(file_path, read_only=True)
            integrity_check["can_open"] = True
            integrity_check["sheet_count"] = len(workbook.sheetnames)
            
            # Verifica se consegue acessar as folhas
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                # Tenta acessar a primeira célula para verificar se a folha está acessível
                for row in sheet.iter_rows(min_row=1, max_row=1, max_col=1):
                    for cell in row:
                        _ = cell.value
                        break
                    break
            
        except InvalidFileException:
            integrity_check["is_corrupt"] = True
            integrity_check["error_message"] = "Arquivo Excel corrompido"
        except Exception as e:
            integrity_check["is_corrupt"] = True
            integrity_check["error_message"] = f"Erro ao verificar integridade: {str(e)}"
        
        return integrity_check